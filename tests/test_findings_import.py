"""Excel AI-review import: deterministic checks aren't duplicated, and importing an AI
review completes the review by auto-approving clean tables.

At project creation the deterministic structural checks (FMT-*/XOUT-*) are already written.
A round-tripped export carries those same rows, so re-importing must SKIP them rather than
add a second copy. Importing also counts as review completion: every 'Not Reviewed' output
with no finding flips to Auto-approved. Self-contained fixture so this file is byte-identical
across both editions.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook

import db
import export


@pytest.fixture()
def iso_db(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DATA_DIR", str(tmp_path))
    monkeypatch.setattr(db, "DB_PATH", str(tmp_path / "app.db"))
    if hasattr(db._local, "conn"):
        del db._local.conn
    db.init()
    yield tmp_path
    if hasattr(db._local, "conn"):
        del db._local.conn


def _project(name="p"):
    return db.insert("project", compound="C", study="S", name=name, edition_label="",
                     created_at=db.now_iso())


def _output(pid, label, seq=0, status="Not Reviewed"):
    did = db.insert("document", project_id=pid, role="delivery", filename="d.pdf",
                    path="d.pdf", n_pages=3, edition="2025")
    return db.insert("output", project_id=pid, document_id=did, seq=seq, output_type="Table",
                     number=label.split()[-1], label=label, title="T",
                     page_start=1, page_end=1, status=status)


def _structural_finding(pid, oid, check_id="FMT-010", label="Table 1"):
    return db.insert("finding", project_id=pid, output_id=oid, run_id=None,
                     check_id=check_id, severity="low", risk="Low",
                     message="blank page detected", subjects="[]", numbers="[]",
                     page=1, printed_page=None, pages_total=None, section=None,
                     row_kind=None, signature="", state="pending", badge="",
                     phase="structural", affected="[]")


def _sheet(rows):
    """rows: list of [Output, Check, Risk, Finding]."""
    wb = Workbook(); ws = wb.active
    ws.append(["Output", "Check", "Risk", "Finding"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_deterministic_rows_are_skipped_not_duplicated(iso_db):
    pid = _project()
    oid = _output(pid, "Table 1")
    _structural_finding(pid, oid, "FMT-010")   # already present from "creation"

    data = _sheet([
        ["Table 1", "FMT-010", "Low", "blank page detected"],      # deterministic dup -> skip
        ["Table 1", "AIW-2.1", "High", "Group-N sum mismatch"],    # real AI finding -> import
    ])
    summary = export.import_findings_xlsx(pid, data)

    assert summary["imported"] == 1
    assert summary["skipped_deterministic"] == 1

    # Exactly one FMT-010 survives (the creation-time one), not two.
    fmt = db.query("SELECT * FROM finding WHERE project_id=? AND check_id LIKE 'FMT%'", (pid,))
    assert len(fmt) == 1
    assert fmt[0]["phase"] == "structural"
    ai = db.query("SELECT * FROM finding WHERE project_id=? AND check_id='AIW-2.1'", (pid,))
    assert len(ai) == 1 and ai[0]["phase"] == "imported"


def test_findings_export_keeps_formula_leading_text_as_strings(iso_db):
    pid = _project(name="=1+1")
    oid = _output(pid, "@formula-like-table")
    fid = _structural_finding(pid, oid, "FMT-010", "@formula-like-table")
    db.execute(
        "UPDATE finding SET message=?, section=? WHERE id=?",
        ('=HYPERLINK("https://invalid.example", "click")', "+formula-like-section", fid),
    )
    db.execute("UPDATE document SET filename=? WHERE project_id=?", ("-formula-like.pdf", pid))

    wb = load_workbook(io.BytesIO(export.findings_xlsx(pid)), data_only=False)
    summary_project = wb["Summary"]["B2"]
    assert summary_project.data_type == "s"
    assert summary_project.value == "'=1+1"

    ws = wb["AI Findings"]
    columns = {cell.value: cell.column for cell in ws[2]}
    for header, expected_prefix in (
        ("Output", "'@"),
        ("Source File", "'-"),
        ("Section", "'+"),
        ("Finding", "'="),
    ):
        cell = ws.cell(row=3, column=columns[header])
        assert cell.data_type == "s"
        assert cell.value.startswith(expected_prefix)


def test_import_auto_approves_clean_tables(iso_db):
    pid = _project()
    o1 = _output(pid, "Table 1", seq=0)
    o2 = _output(pid, "Table 2", seq=1)   # will stay clean

    summary = export.import_findings_xlsx(
        pid, _sheet([["Table 1", "AIW-2.1", "High", "sum mismatch"]]))

    assert summary["auto_approved"] == 1
    assert db.one("SELECT status FROM output WHERE id=?", (o1,))["status"] == "Not Reviewed"
    assert db.one("SELECT status FROM output WHERE id=?", (o2,))["status"] == "Auto-approved"


def test_table_with_deterministic_finding_is_not_approved(iso_db):
    # A table flagged only by a creation-time structural check is NOT clean: the skipped
    # deterministic row keeps its finding attached, so the table must not auto-approve.
    pid = _project()
    oid = _output(pid, "Table 1")
    _structural_finding(pid, oid, "FMT-010")

    summary = export.import_findings_xlsx(
        pid, _sheet([["Table 1", "FMT-010", "Low", "blank page detected"]]))

    assert summary["imported"] == 0 and summary["skipped_deterministic"] == 1
    assert summary["auto_approved"] == 0
    assert db.one("SELECT status FROM output WHERE id=?", (oid,))["status"] == "Not Reviewed"


def test_import_never_overwrites_human_status(iso_db):
    pid = _project()
    oid = _output(pid, "Table 1", status="Rejected")   # human already set a status

    summary = export.import_findings_xlsx(pid, _sheet([]))   # empty -> nothing to import

    assert summary["auto_approved"] == 0
    assert db.one("SELECT status FROM output WHERE id=?", (oid,))["status"] == "Rejected"


@pytest.mark.parametrize(
    "limit_name,limit_value,needle",
    [
        ("_MAX_XLSX_ENTRIES", 1, "too many archive entries"),
        ("_MAX_XLSX_ENTRY_BYTES", 1, "entry is too large"),
        ("_MAX_XLSX_TOTAL_BYTES", 1, "expands beyond"),
    ],
)
def test_xlsx_container_limits_apply_before_openpyxl(
    iso_db, monkeypatch, limit_name, limit_value, needle,
):
    pid = _project()
    _output(pid, "Table 1")
    monkeypatch.setattr(export, limit_name, limit_value)
    with pytest.raises(ValueError, match=needle):
        export.import_findings_xlsx(
            pid, _sheet([["Table 1", "AIW-2.1", "High", "sum mismatch"]]),
        )
    assert db.query("SELECT id FROM ai_run WHERE project_id=?", (pid,)) == []


def test_xlsx_compression_ratio_and_sheet_shape_limits(iso_db, monkeypatch):
    pid = _project()
    _output(pid, "Table 1")
    workbook = _sheet([["Table 1", "AIW-2.1", "High", "x" * 5000]])

    monkeypatch.setattr(export, "_XLSX_RATIO_CHECK_MIN_BYTES", 1)
    monkeypatch.setattr(export, "_MAX_XLSX_COMPRESSION_RATIO", 1)
    with pytest.raises(ValueError, match="unsafe compression ratio"):
        export.import_findings_xlsx(pid, workbook)

    monkeypatch.setattr(export, "_MAX_XLSX_COMPRESSION_RATIO", 10_000)
    monkeypatch.setattr(export, "_MAX_IMPORT_ROWS", 1)
    with pytest.raises(ValueError, match="row / .*column"):
        export.import_findings_xlsx(pid, workbook)


def test_legacy_approved_migration_splits_auto_vs_manual(iso_db):
    # Legacy rows only ever stored the pre-split 'Approved'. Re-running init() must reclassify
    # them exactly as a fresh run would: a clean table (no finding) becomes 'Auto-approved',
    # a table that still carries a finding becomes 'Manually approved'.
    pid = _project()
    clean = _output(pid, "Table 1", status="Approved")          # no finding → AI clean pass
    flagged = _output(pid, "Table 2", seq=1, status="Approved")  # human approved despite a finding
    _structural_finding(pid, flagged, "FMT-010", "Table 2")

    db.init()   # idempotent; runs the status migration

    assert db.one("SELECT status FROM output WHERE id=?", (clean,))["status"] == "Auto-approved"
    assert db.one("SELECT status FROM output WHERE id=?", (flagged,))["status"] == "Manually approved"
