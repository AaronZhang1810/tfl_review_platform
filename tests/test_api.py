"""API smoke tests with an isolated temp database (no real app.db touched).

Covers project create/list/delete, the path-traversal defense end-to-end, and crash-recovery of interrupted AI runs."""

import io
import os

import pytest
from pypdf import PdfWriter


@pytest.fixture()
def client(tmp_path, monkeypatch):
    import db
    # Point the DB + uploads at a temp dir and reset the thread-local connection.
    monkeypatch.setattr(db, "DATA_DIR", str(tmp_path))
    monkeypatch.setattr(db, "DB_PATH", str(tmp_path / "test.db"))
    if hasattr(db._local, "conn"):
        del db._local.conn
    import main
    monkeypatch.setattr(main, "UPLOADS", str(tmp_path / "uploads"))
    from fastapi.testclient import TestClient
    with TestClient(main.app) as c:
        yield c


def _blank_pdf(pages: int = 1) -> bytes:
    w = PdfWriter()
    for _ in range(pages):
        w.add_blank_page(width=200, height=200)
    b = io.BytesIO()
    w.write(b)
    return b.getvalue()


def _numbering_gap_pdf() -> bytes:
    # Two bookmarked outputs, "Table 1" and "Table 3" — the missing Table 2 is a deterministic numbering gap (XOUT-001) the indexer + structural checks must surface.
    w = PdfWriter()
    w.add_blank_page(width=200, height=200)
    w.add_blank_page(width=200, height=200)
    w.add_outline_item("Table 1  Demographics", 0)
    w.add_outline_item("Table 3  Adverse events", 1)
    b = io.BytesIO()
    w.write(b)
    return b.getvalue()


def test_create_project_runs_structural_checks(client):
    # Creation must run the deterministic checks (no AI key) and report + persist them.
    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "struct"},
                    files=[("delivery", ("d.pdf", _numbering_gap_pdf(), "application/pdf"))])
    assert r.status_code == 200, r.text
    body = r.json()
    pid = body["id"]
    assert body["n_structural"] >= 1, body

    findings = client.get(f"/api/projects/{pid}/findings").json()
    structural = [f for f in findings if f.get("phase") == "structural"]
    assert any(f["check_id"] == "XOUT-001" for f in structural), structural
    # No AI run was created, so the AI-last-run endpoint still reports none.
    assert client.get(f"/api/projects/{pid}/ai-last-run").json().get("none") is True


def test_create_list_delete_project(client):
    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "smoke"},
                    files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))])
    assert r.status_code == 200, r.text
    pid = r.json()["id"]

    projects = client.get("/api/projects").json()
    assert any(p["id"] == pid for p in projects)

    assert client.delete(f"/api/projects/{pid}").status_code == 200
    assert not any(p["id"] == pid for p in client.get("/api/projects").json())


def test_upload_filename_is_sanitized(client):
    # A traversal filename must be stored under the project folder, basename only.
    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "trav"},
                    files=[("delivery", ("../../pwned.pdf", _blank_pdf(), "application/pdf"))])
    assert r.status_code == 200, r.text
    pid = r.json()["id"]
    docs = client.get(f"/api/projects/{pid}").json()["documents"]
    assert docs[0]["filename"] == "pwned.pdf"
    assert ".." not in docs[0]["filename"]


def test_upload_filename_collisions_do_not_overwrite_documents(client):
    response = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "collisions", "main_index": 0},
        files=[
            ("delivery", ("same.pdf", _blank_pdf(), "application/pdf")),
            ("delivery", ("same.pdf", _blank_pdf(2), "application/pdf")),
        ],
    )
    assert response.status_code == 200, response.text
    import db
    documents = db.query(
        "SELECT filename, path, n_pages FROM document WHERE project_id=? ORDER BY id",
        (response.json()["id"],),
    )
    assert [row["n_pages"] for row in documents] == [1, 2]
    assert len({row["filename"].casefold() for row in documents}) == 2
    assert len({row["path"].casefold() for row in documents}) == 2


def test_malformed_or_excessive_pdf_is_rejected_and_cleaned(client, monkeypatch):
    malformed = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "bad-pdf"},
        files=[("delivery", ("fake.pdf", b"not a PDF", "application/pdf"))],
    )
    assert malformed.status_code == 400

    import project_io
    monkeypatch.setattr(project_io, "_MAX_PDF_PAGES", 1)
    excessive = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "too-many-pages"},
        files=[("delivery", ("two.pdf", _blank_pdf(2), "application/pdf"))],
    )
    assert excessive.status_code == 400
    assert client.get("/api/projects").json() == []


def test_document_upload_limit_is_fail_closed_and_cleans_partial_project(client, monkeypatch):
    import main

    pdf = _blank_pdf()
    monkeypatch.setattr(main, "MAX_DOCUMENT_BYTES", len(pdf))
    response = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "oversize"},
        files=[("delivery", ("too-big.pdf", pdf + b"x", "application/pdf"))],
    )

    assert response.status_code == 413
    assert client.get("/api/projects").json() == []
    upload_root = os.path.join(main.UPLOADS)
    assert not os.path.exists(upload_root) or os.listdir(upload_root) == []


def test_document_upload_at_exact_limit_is_accepted(client, monkeypatch):
    import main

    pdf = _blank_pdf()
    monkeypatch.setattr(main, "MAX_DOCUMENT_BYTES", len(pdf))
    response = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "boundary"},
        files=[("delivery", ("boundary.pdf", pdf, "application/pdf"))],
    )

    assert response.status_code == 200, response.text


def test_declared_request_size_limit_rejects_before_mutation(client, monkeypatch):
    import main

    monkeypatch.setattr(main, "MAX_REQUEST_BYTES", 1)
    response = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "request-limit"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )

    assert response.status_code == 413
    assert client.get("/api/projects").json() == []


def test_project_bundle_upload_limit_returns_413(client, monkeypatch):
    import main

    monkeypatch.setattr(main, "MAX_BUNDLE_BYTES", 4)
    response = client.post(
        "/api/projects/import",
        data={"mode": "new"},
        files=[("file", ("project.zip", b"12345", "application/zip"))],
    )

    assert response.status_code == 413


def test_workbook_upload_limits_return_413(client, monkeypatch):
    import main

    created = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "sheet-limit"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )
    pid = created.json()["id"]
    monkeypatch.setattr(main, "MAX_SHEET_BYTES", 4)

    for path in (
        f"/api/projects/{pid}/import-comments",
        f"/api/projects/{pid}/import-findings",
    ):
        response = client.post(
            path,
            files=[("file", ("sheet.xlsx", b"12345", "application/octet-stream"))],
        )
        assert response.status_code == 413, (path, response.text)


def test_audit_trail_records_mutations(client):
    # Create a project + output, change status, and confirm the audit log captured it.
    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "audit"},
                    files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))])
    pid = r.json()["id"]
    import db
    oid = db.insert("output", project_id=pid, document_id=1, seq=0, output_type="Table",
                    number="1", label="Table 1", title="T", page_start=1, page_end=1)
    client.post(f"/api/outputs/{oid}/status",
                data={"status": "Manually approved", "actor": "Dr X"})
    log = client.get(f"/api/projects/{pid}/audit").json()
    actions = {e["action"] for e in log}
    assert "project.create" in actions
    assert "status.set" in actions
    assert any(e["actor"] == "Dr X" and e["action"] == "status.set" for e in log)


def test_status_endpoint_rejects_untrusted_markup(client):
    import db

    created = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "status-safety"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )
    pid = created.json()["id"]
    document_id = db.one("SELECT id FROM document WHERE project_id=?", (pid,))["id"]
    oid = db.insert(
        "output", project_id=pid, document_id=document_id, seq=0,
        output_type="Table", number="1", label="Table 1", title="T",
        page_start=1, page_end=1,
    )

    response = client.post(
        f"/api/outputs/{oid}/status",
        data={"status": '\"><img src=x onerror=alert(1)>'},
    )

    assert response.status_code == 400
    assert db.one("SELECT status FROM output WHERE id=?", (oid,))["status"] == "Not Reviewed"


def test_untrusted_host_and_cross_origin_mutation_are_rejected(client):
    bad_host = client.get("/api/projects", headers={"Host": "attacker.example"})
    assert bad_host.status_code == 400

    bad_origin = client.post(
        "/api/projects/import",
        headers={"Origin": "https://attacker.example"},
        files=[("file", ("project.zip", b"not-a-zip", "application/zip"))],
    )
    assert bad_origin.status_code == 403

    wrong_port = client.post(
        "/api/projects/import",
        headers={"Origin": "http://testserver:81"},
        files=[("file", ("project.zip", b"not-a-zip", "application/zip"))],
    )
    assert wrong_port.status_code == 403

    fetch_metadata = client.post(
        "/api/projects/import",
        headers={"Sec-Fetch-Site": "cross-site"},
        files=[("file", ("project.zip", b"not-a-zip", "application/zip"))],
    )
    assert fetch_metadata.status_code == 403


def test_api_is_no_store_and_content_disposition_is_sanitized(client):
    import db

    created = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "headers"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )
    pid = created.json()["id"]
    document = db.one("SELECT id FROM document WHERE project_id=?", (pid,))
    oid = db.insert(
        "output", project_id=pid, document_id=document["id"], seq=0,
        output_type="Table", number="1", label='bad"\r\nX-Injected: yes', title="T",
        page_start=1, page_end=1,
    )

    listing = client.get("/api/projects")
    assert listing.headers["cache-control"] == "no-store"
    script_policy = next(
        directive for directive in listing.headers["content-security-policy"].split("; ")
        if directive.startswith("script-src ")
    )
    assert "'unsafe-inline'" not in script_policy
    assert "'sha256-" in script_policy
    clip = client.get(f"/api/tlf-clip?output_id={oid}")
    assert clip.status_code == 200
    assert clip.headers["cache-control"] == "no-store"
    assert "\r" not in clip.headers["content-disposition"]
    assert "\n" not in clip.headers["content-disposition"]
    assert "x-injected" not in {key.lower() for key in clip.headers}


def test_annotation_payload_is_schema_validated(client):
    import db

    created = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "annotation"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )
    pid = created.json()["id"]
    document = db.one("SELECT id FROM document WHERE project_id=?", (pid,))
    oid = db.insert(
        "output", project_id=pid, document_id=document["id"], seq=0,
        output_type="Table", number="1", label="Table 1", title="T",
        page_start=1, page_end=1,
    )

    hostile = client.post(
        f"/api/outputs/{oid}/annotations",
        data={"kind": "rect", "page": 1, "geom_json": '{"x":0,"y":0,"w":2,"h":1}'},
    )
    assert hostile.status_code == 400
    valid = client.post(
        f"/api/outputs/{oid}/annotations",
        data={"kind": "rect", "page": 1,
              "geom_json": '{"x":0.1,"y":0.2,"w":0.3,"h":0.4,"color":"#abcdef"}'},
    )
    assert valid.status_code == 200
    stored = db.one("SELECT geom_json FROM annotation WHERE id=?", (valid.json()["id"],))
    assert stored["geom_json"] == '{"x":0.1,"y":0.2,"w":0.3,"h":0.4,"color":"#abcdef"}'


def test_project_delete_removes_non_fk_audit_and_review_history(client):
    import db

    created = client.post(
        "/api/projects",
        data={"compound": "C", "study": "S", "name": "delete-history"},
        files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))],
    )
    pid = created.json()["id"]
    db.insert("review_log", ts=db.now_iso(), reviewer="R", action="comment_add",
              project_id=pid, output_label="", check_id="", checklist_item="", risk="",
              message="", numbers="[]", subjects="[]", page=None, printed_page=None,
              section=None, comment_text="text", finding_signature="")
    assert db.query("SELECT id FROM audit_log WHERE project_id=?", (pid,))
    assert db.query("SELECT id FROM review_log WHERE project_id=?", (pid,))

    assert client.delete(f"/api/projects/{pid}").status_code == 200
    assert db.query("SELECT id FROM audit_log WHERE project_id=?", (pid,)) == []
    assert db.query("SELECT id FROM review_log WHERE project_id=?", (pid,)) == []
    assert client.delete(f"/api/projects/{pid}").status_code == 404


def test_recover_stale_runs(client):
    import db
    db.insert("project", compound="C", study="S", name="r", edition_label="", created_at=db.now_iso())
    rid = db.insert("ai_run", project_id=1, kind="fresh", started_at=db.now_iso(),
                    finished_at=None, summary_json="{}")
    assert db.recover_stale_runs() >= 1
    row = db.one("SELECT finished_at, summary_json FROM ai_run WHERE id=?", (rid,))
    assert row["finished_at"] is not None
    assert "interrupted" in row["summary_json"]


def test_ai_run_returns_conflict_while_process_lease_is_held(client, monkeypatch):
    import main
    import runner

    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "lease"},
                    files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))])
    pid = r.json()["id"]
    monkeypatch.setattr(main.ai_client, "available", lambda: True)
    token = runner._acquire_project_lease(pid)
    try:
        conflict = client.post(f"/api/projects/{pid}/ai-run",
                               data={"model": "", "effort": ""})
        assert conflict.status_code == 409
        assert "already running" in conflict.text
    finally:
        runner._release_project_lease(pid, token)


def test_comments_expose_page_and_roundtrip_import(client):
    # Comments list returns the table page; an exported sheet re-imports body edits.
    import io
    import db
    from openpyxl import load_workbook
    r = client.post("/api/projects",
                    data={"compound": "C", "study": "S", "name": "rt"},
                    files=[("delivery", ("d.pdf", _blank_pdf(), "application/pdf"))])
    pid = r.json()["id"]
    oid = db.insert("output", project_id=pid, document_id=1, seq=0, output_type="Table",
                    number="1", label="Table 1", title="T", page_start=5, page_end=5)
    client.post(f"/api/outputs/{oid}/comments", data={"title": "t", "body": "original"})

    comments = client.get(f"/api/projects/{pid}/comments").json()
    assert comments and comments[0]["page"] == 5

    xlsx = client.get(f"/api/projects/{pid}/export/comments.xlsx").content
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    body_col = headers.index("Comment") + 1
    ws.cell(row=2, column=body_col, value="edited via excel")
    buf = io.BytesIO()
    wb.save(buf)

    imp = client.post(f"/api/projects/{pid}/import-comments",
                      files=[("file", ("comments.xlsx", buf.getvalue(),
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
    assert imp.status_code == 200, imp.text
    assert imp.json()["updated"] >= 1

    comments = client.get(f"/api/projects/{pid}/comments").json()
    assert any(c["body"] == "edited via excel" for c in comments)
