"""Comment export→edit→re-import round-trip (export.comments_xlsx / import_comments_xlsx).

New schema: ID | Table | Comment | Reply to | Resolved. `(ID, Table)` is the identity — a row
whose (ID, Table) already exists REPLACES that comment; a new (ID, Table) on an existing Table
CREATES one; "Reply to" points at another comment's per-Table ID. The whole sheet is validated
up front and applied atomically, so any bad row aborts the import with a ValueError and writes
nothing. Self-contained fixture so this file is byte-identical across both editions.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook

import db
import export


@pytest.fixture()
def iso_db(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DATA_DIR", str(tmp_path))
    monkeypatch.setattr(db, "DB_PATH", str(tmp_path / "app.db"))
    if hasattr(db._local, "conn"):
        del db._local.conn
    db.init()
    yield tmp_path
    if hasattr(db._local, "conn"):
        del db._local.conn


def _project(name="p"):
    return db.insert("project", compound="C", study="S", name=name, edition_label="",
                     created_at=db.now_iso())


def _seed_output(pid, label="Table 1", page_start=7):
    did = db.insert("document", project_id=pid, role="delivery", filename="d.pdf",
                    path="d.pdf", n_pages=9, edition="2025")
    return db.insert("output", project_id=pid, document_id=did, seq=0, output_type="Table",
                     number=label.split()[-1], label=label, title="Summary",
                     page_start=page_start, page_end=9)


def _comment(pid, oid, num, body, parent_id=None, resolved=0, source="manual"):
    return db.insert("comment", project_id=pid, output_id=oid, num=num, title="", body=body,
                     source=source, finding_id=None, parent_id=parent_id, resolved=resolved,
                     created_at=db.now_iso())


def _sheet(rows):
    """rows: list of [ID, Table, Comment, Reply to, Resolved]."""
    wb = Workbook(); ws = wb.active; ws.title = "Review Comments"
    ws.append(["ID", "Table", "Comment", "Reply to", "Resolved"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- export

def test_export_uses_new_schema(iso_db):
    pid = _project()
    oid = _seed_output(pid, "Table 1")
    c1 = _comment(pid, oid, 1, "first comment", resolved=0)
    _comment(pid, oid, 2, "a reply", parent_id=c1, resolved=1, source="reply")

    ws = load_workbook(io.BytesIO(export.comments_xlsx(pid))).active
    assert [c.value for c in ws[1]] == ["ID", "Table", "Comment", "Reply to", "Resolved"]

    by_id = {r[0].value: [c.value for c in r] for r in ws.iter_rows(min_row=2)}
    assert by_id[1] == [1, "Table 1", "first comment", None, "No"]
    assert by_id[2] == [2, "Table 1", "a reply", 1, "Yes"]


def test_comment_export_keeps_formula_leading_text_as_strings(iso_db):
    pid = _project()
    oid = _seed_output(pid, "@formula-like-table")
    _comment(pid, oid, 1, '=HYPERLINK("https://invalid.example", "click")')

    ws = load_workbook(io.BytesIO(export.comments_xlsx(pid)), data_only=False).active
    table_cell = ws["B2"]
    comment_cell = ws["C2"]
    assert table_cell.data_type == "s"
    assert table_cell.value == "'@formula-like-table"
    assert comment_cell.data_type == "s"
    assert comment_cell.value.startswith("'=HYPERLINK")


# --------------------------------------------------------------------------- round-trip

def test_roundtrip_replace_create_and_reply(iso_db):
    pid = _project()
    oid = _seed_output(pid, "Table 1")
    c1 = _comment(pid, oid, 1, "old body", resolved=0)

    summary = export.import_comments_xlsx(pid, _sheet([
        [1, "Table 1", "edited body", "", "yes"],       # replaces c1
        [2, "Table 1", "new top-level", "", ""],         # creates num 2
        [3, "Table 1", "a reply to 1", 1, "no"],         # creates num 3, replying to 1
    ]))
    assert summary["updated"] == 1
    assert summary["created"] == 2

    replaced = db.one("SELECT * FROM comment WHERE id=?", (c1,))
    assert replaced["body"] == "edited body"
    assert replaced["resolved"] == 1

    reply = db.one("SELECT * FROM comment WHERE output_id=? AND num=3", (oid,))
    assert reply["body"] == "a reply to 1"
    assert reply["parent_id"] == c1                      # reply-to num 1 → c1's global id
    assert db.query("SELECT id FROM comment WHERE project_id=?", (pid,)).__len__() == 3


def test_reply_to_a_row_created_in_the_same_sheet(iso_db):
    pid = _project()
    oid = _seed_output(pid, "Table 1")
    export.import_comments_xlsx(pid, _sheet([
        [1, "Table 1", "parent", "", ""],
        [2, "Table 1", "child", 1, ""],
    ]))
    parent = db.one("SELECT id FROM comment WHERE output_id=? AND num=1", (oid,))
    child = db.one("SELECT * FROM comment WHERE output_id=? AND num=2", (oid,))
    assert child["parent_id"] == parent["id"]


def test_num_is_scoped_per_table(iso_db):
    # The same ID may exist in two different Tables — (ID, Table) is the key, not ID alone.
    pid = _project()
    o1 = _seed_output(pid, "Table 1")
    o2 = _seed_output(pid, "Table 2")
    export.import_comments_xlsx(pid, _sheet([
        [1, "Table 1", "in table 1", "", ""],
        [1, "Table 2", "in table 2", "", ""],
    ]))
    assert db.one("SELECT body FROM comment WHERE output_id=? AND num=1", (o1,))["body"] == "in table 1"
    assert db.one("SELECT body FROM comment WHERE output_id=? AND num=1", (o2,))["body"] == "in table 2"


# --------------------------------------------------------------------------- validation

@pytest.mark.parametrize("rows, needle", [
    ([[1, "Table 1", "", "", ""]], "must all be filled in"),           # empty Comment
    ([["", "Table 1", "body", "", ""]], "must all be filled in"),      # empty ID
    ([[1, "", "body", "", ""]], "must all be filled in"),              # empty Table
    ([["x", "Table 1", "body", "", ""]], "positive whole number"),     # non-integer ID
    ([[0, "Table 1", "body", "", ""]], "positive whole number"),       # non-positive ID
    ([[1, "No Such Table", "body", "", ""]], "does not exist"),        # unknown Table
    ([[1, "Table 1", "body", "", "maybe"]], "Resolved must be"),       # bad Resolved
    ([[1, "Table 1", "body", "abc", ""]], "'Reply to' must be"),       # non-integer Reply to
    ([[1, "Table 1", "body", 1, ""]], "cannot reply to itself"),       # self reply
    ([[1, "Table 1", "body", 99, ""]], "does not match any comment"),  # dangling reply
    ([[1, "Table 1", "a", "", ""], [1, "Table 1", "b", "", ""]], "duplicate"),  # dup (ID,Table)
])
def test_import_validation_errors(iso_db, rows, needle):
    pid = _project()
    _seed_output(pid, "Table 1")
    _seed_output(pid, "Table 2")
    with pytest.raises(ValueError) as ei:
        export.import_comments_xlsx(pid, _sheet(rows))
    assert needle in str(ei.value)


def test_failed_import_writes_nothing(iso_db):
    # A single bad row aborts the whole sheet — the good rows before it are not written.
    pid = _project()
    _seed_output(pid, "Table 1")
    with pytest.raises(ValueError):
        export.import_comments_xlsx(pid, _sheet([
            [1, "Table 1", "would be created", "", ""],
            [2, "Table 1", "", "", ""],   # empty Comment → abort
        ]))
    assert db.query("SELECT id FROM comment WHERE project_id=?", (pid,)) == []


def test_import_rejects_wrong_sheet(iso_db):
    pid = _project()
    wb = Workbook(); wb.active.append(["Foo", "Bar"])
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(ValueError):
        export.import_comments_xlsx(pid, buf.getvalue())
