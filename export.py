"""Export helpers: comment log and AI-findings report to Excel."""

from __future__ import annotations

import io
import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import ai_review
import checks

# One two-tier scale. _TIER maps any stored value — current (high/low) or legacy (critical/major/minor) — to a High-first sort rank so old exports still order right.
_TIER = {"high": 0, "critical": 0, "major": 0, "low": 1, "minor": 1, "medium": 1}
_RISK_FILL = {"High": "F8D7DA", "Low": "E8EBEF"}
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _excel_value(value):
    """Keep untrusted text from becoming a formula in an exported workbook.

Excel-compatible applications interpret several leading characters as formulas or commands. Prefixing an apostrophe keeps the cell textual (the apostrophe is normally not displayed by spreadsheet applications). Numeric values are retained as numbers so sorting and summaries continue to work."""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _tier_label(r: dict) -> str:
    """The finding's single tier as High/Low, from risk (preferred) or legacy severity."""
    v = (r.get("risk") or r.get("severity") or "").lower()
    return "High" if _TIER.get(v, 1) == 0 else "Low"
_STATUS_FILL = {"posted": "D9EAD3", "rejected": "EAECEE", "resolved": "D9EAD3", "pending": "FFF3CD"}


def comments_xlsx(project_id: int) -> bytes:
    # ID | Table | Comment | Reply to | Resolved. `num` is the per-Table comment ID (see db.py); (ID, Table) is the round-trip key. "Reply to" carries the parent comment's own num (empty for a top-level comment); resolved reads back as Yes/No.
    rows = db.query(
        """SELECT c.num, o.label AS tbl, c.body, c.resolved, p.num AS reply_to
           FROM comment c LEFT JOIN output o ON o.id = c.output_id
                          LEFT JOIN comment p ON p.id = c.parent_id
           WHERE c.project_id = ? ORDER BY o.label, c.num""",
        (project_id,),
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Comments"
    headers = ["ID", "Table", "Comment", "Reply to", "Resolved"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F3B5B")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([
            r["num"], _excel_value(r["tbl"] or ""), _excel_value(r["body"] or ""),
            r["reply_to"] or "",
            "Yes" if r["resolved"] else "No",
        ])
    widths = [8, 18, 70, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")   # Comment
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# AI findings report
# --------------------------------------------------------------------------- #

def _check_descriptions() -> dict:
    """{check_id -> human description} built from the checklist config (title + guidance), plus the structural/helper findings that aren't checklist items."""
    try:
        idx = ai_review.checklist_index(ai_review.load_config())
        out = {}
        for cid, it in idx.items():
            title, guidance = it.get("title", ""), it.get("guidance", "")
            out[cid] = f"{title} — {guidance}" if guidance else title
        out.setdefault("FMT-002", "Row is missing an N where one is expected")
        return out
    except Exception:
        return {}


def _scope(check_id: str, output_id) -> str:
    """Prefix-based scope. AIV-* / XOUT-020 compare against the prior edition; AIX-* / XOUT-001 / findings without an owning output are cross-output; the rest are within a single file."""
    cid = check_id or ""
    if cid.startswith("AIV-") or cid == "XOUT-020":
        return "Cross-file (vs prior edition)"
    if cid.startswith("AIX-") or cid == "XOUT-001" or output_id is None:
        return "Cross-output"
    return "Within file"


def findings_xlsx(project_id: int) -> bytes:
    proj = db.one("SELECT * FROM project WHERE id=?", (project_id,)) or {}
    rows = db.query(
        """SELECT f.*, o.label AS output_label, o.title AS output_title,
                  d.filename AS file
           FROM finding f
           LEFT JOIN output o ON o.id=f.output_id
           LEFT JOIN document d ON d.id=o.document_id
           WHERE f.project_id=?""", (project_id,))
    # linked comments per finding (posted findings become comments)
    clinks: dict[int, list[str]] = {}
    for c in db.query("SELECT finding_id, author, body FROM comment WHERE project_id=? AND finding_id IS NOT NULL",
                      (project_id,)):
        clinks.setdefault(c["finding_id"], []).append(f"{c.get('author') or 'Reviewer'}: {c['body']}")

    descs = _check_descriptions()
    for r in rows:                       # single tier, legacy-safe, used by both sheets
        r["_tier"] = _tier_label(r)
    rows.sort(key=lambda r: (0 if r["_tier"] == "High" else 1, r["output_label"] or "~",
                             0 if (r.get("row_kind") or "study") != "aggregate" else 1,
                             r["check_id"]))

    wb = Workbook()
    _findings_sheet(wb.active, proj, rows, clinks, descs)
    _summary_sheet(wb.create_sheet("Summary", 0), proj, rows)
    wb.active = wb["AI Findings"] if "AI Findings" in wb.sheetnames else wb.worksheets[0]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_COLS = [
    ("Output", 14, "output_label"),
    ("Output Title", 40, "output_title"),
    ("Check", 11, "check_id"),
    ("Check Description", 34, "_desc"),
    ("Source File", 22, "file"),
    ("Scope", 15, "_scope"),
    ("Risk", 9, "_tier"),
    ("Status", 11, "state"),
    ("Page", 7, "page"),
    ("Subtable", 10, "_subtable"),
    ("Section", 16, "section"),
    ("Row type", 10, "row_kind"),
    ("Subjects", 18, "_subjects"),
    ("Finding", 70, "message"),
    ("Reviewer Comments", 40, "_comments"),
]

_COL_KEYS = [k for _, _, k in _COLS]


def _findings_sheet(ws, proj, rows, clinks, descs):
    ws.title = "AI Findings"
    thin = Side(style="thin", color="D0D4DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = f"AI Review Findings — {proj.get('name', '')}".strip(" —")
    ws.append([_excel_value(title)])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLS))
    ws["A1"].font = Font(bold=True, size=13, color="1F3B5B")

    hrow = 2
    ws.append([c[0] for c in _COLS])
    head_fill = PatternFill("solid", fgColor="1F3B5B")
    for c in ws[hrow]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border

    for r in rows:
        vals = {
            **r,
            "_desc": descs.get(r["check_id"], ""),
            "_scope": _scope(r["check_id"], r["output_id"]),
            "_subjects": ", ".join(db.loads(r.get("subjects"), []) or []),
            "_comments": "\n".join(clinks.get(r["id"], [])),
            "output_label": r["output_label"] or "(cross-output)",
            "_subtable": (f"p{r['printed_page']}/{r['pages_total']}"
                          if r.get("printed_page") and r.get("pages_total") else ""),
        }
        ws.append([_excel_value(vals.get(key, "")) for _, _, key in _COLS])
        row = ws[ws.max_row]
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        # color the risk + status cells (index by key so column order can move)
        risk_cell = row[_COL_KEYS.index("_tier")]
        st_cell = row[_COL_KEYS.index("state")]
        risk_cell.fill = PatternFill("solid", fgColor=_RISK_FILL.get(vals["_tier"], "FFFFFF"))
        risk_cell.font = Font(bold=True)
        st_cell.fill = PatternFill("solid", fgColor=_STATUS_FILL.get(r["state"], "FFFFFF"))

    for i, (_, w, _k) in enumerate(_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(_COLS))}{ws.max_row}"


def _summary_sheet(ws, proj, rows):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.append(["AI Review Summary"])
    ws["A1"].font = Font(bold=True, size=13, color="1F3B5B")
    ws.append(["Project", _excel_value(proj.get("name", ""))])
    ws.append(["Total findings", len(rows)])
    ws.append([])

    def block(title, key, order=None):
        ws.append([title]); ws[f"A{ws.max_row}"].font = Font(bold=True)
        counts: dict = {}
        for r in rows:
            counts[r[key]] = counts.get(r[key], 0) + 1
        keys = order or sorted(counts)
        for k in keys:
            if k in counts:
                ws.append([f"  {k}", counts[k]])
        ws.append([])

    block("By risk", "_tier", ["High", "Low"])
    block("By status", "state", ["pending", "posted", "rejected", "resolved"])
    block("By check", "check_id")


# --------------------------------------------------------------------------- #
# Annotated-PDF export: bake reviewer marks + comment threads into one PDF
# --------------------------------------------------------------------------- #

def annotated_pdf(project_id: int) -> bytes:
    import pypdf
    from pypdf.annotations import Rectangle, Highlight, PolyLine, FreeText
    import pdftools

    writer = pypdf.PdfWriter()
    outputs = db.query(
        """SELECT o.*, d.path AS doc_path FROM output o JOIN document d ON d.id=o.document_id
           WHERE o.project_id=? ORDER BY o.document_id, o.seq""", (project_id,))
    annos_by_out = {}
    for a in db.query("""SELECT a.* FROM annotation a JOIN output o ON o.id=a.output_id
                         WHERE o.project_id=?""", (project_id,)):
        annos_by_out.setdefault(a["output_id"], []).append(a)
    comments = db.query(
        """SELECT c.*, o.id AS oid FROM comment c JOIN output o ON o.id=c.output_id
           WHERE c.project_id=? ORDER BY c.created_at""", (project_id,))
    comments_by_out = {}
    for c in comments:
        comments_by_out.setdefault(c["oid"], []).append(c)

    included = 0
    for o in outputs:
        anns = annos_by_out.get(o["id"], [])
        cmts = comments_by_out.get(o["id"], [])
        if not anns and not cmts:
            continue
        included += 1
        # Clip this output's pages into a reader and append to the writer.
        reader = pypdf.PdfReader(io.BytesIO(pdftools.clip(o["doc_path"], o["page_start"], o["page_end"])))
        start_index = len(writer.pages)
        for pg in reader.pages:
            writer.add_page(pg)
        writer.add_outline_item(f"{o['label']} — {o['title']}"[:120], start_index)

        # Bake annotations onto their pages.
        for a in anns:
            try:
                _add_annotation(writer, start_index + (a["page"] - 1), a)
            except Exception:
                pass  # never let one bad mark abort the whole export

        # Append a comments page for this output.
        if cmts:
            w = float(reader.pages[0].mediabox.width)
            h = float(reader.pages[0].mediabox.height)
            writer.add_blank_page(width=w, height=h)
            lines = [f"Review comments — {o['label']}", ""]
            for c in cmts:
                tag = " (AI)" if c["source"] == "ai" else (" (reply)" if c["source"] == "reply" else "")
                lines.append(f"[{c.get('author') or 'Reviewer'}{tag}] {c.get('title') or ''}".rstrip())
                lines.append(f"  {c['body']}")
                lines.append("")
            try:
                writer.add_annotation(page_number=len(writer.pages) - 1,
                                      annotation=FreeText(text="\n".join(lines),
                                                          rect=(36, 36, w - 36, h - 36), font_size="10pt"))
            except Exception:
                pass

    if included == 0:
        writer.add_blank_page(width=612, height=792)
        writer.add_outline_item("No annotations or comments yet", 0)

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def _add_annotation(writer, page_number, a):
    """Convert a normalized annotation to a PDF annotation on the given page."""
    import json as _json
    from pypdf.annotations import Rectangle, Highlight, PolyLine
    from pypdf.generic import ArrayObject, FloatObject
    page = writer.pages[page_number]
    W = float(page.mediabox.width); H = float(page.mediabox.height)
    g = _json.loads(a["geom_json"] or "{}")
    color = (g.get("color") or "#ffd54a").lstrip("#")

    if a["kind"] in ("rect", "highlight"):
        x1 = g["x"] * W; x2 = (g["x"] + g["w"]) * W
        y2 = H - g["y"] * H; y1 = H - (g["y"] + g["h"]) * H          # flip: PDF origin is bottom-left
        if a["kind"] == "highlight":
            quad = ArrayObject([FloatObject(v) for v in (x1, y2, x2, y2, x1, y1, x2, y1)])
            ann = Highlight(rect=(x1, y1, x2, y2), quad_points=quad, highlight_color=color)
        else:
            ann = Rectangle(rect=(x1, y1, x2, y2), interior_color=None)
        writer.add_annotation(page_number=page_number, annotation=ann)
    elif a["kind"] == "freehand" and g.get("pts"):
        verts = [(px * W, H - py * H) for px, py in g["pts"]]
        if len(verts) >= 2:
            writer.add_annotation(page_number=page_number, annotation=PolyLine(vertices=verts))


# --------------------------------------------------------------------------- #
# Import: load findings from a structured Excel file (the reverse of the export above). Lets a reviewer author findings offline, or export → edit → re-import. Columns are matched by HEADER NAME (case-insensitive), so the export's own sheet round-trips and a hand-made sheet works as long as it has a "Finding" column.
# --------------------------------------------------------------------------- #

import re as _re
from openpyxl import load_workbook

def _positive_env_int(name: str, default: int) -> int:
    raw = os.environ.get(name)
    if raw is None:
        return default
    try:
        value = int(raw)
    except ValueError as exc:
        raise RuntimeError(f"{name} must be a positive integer") from exc
    if value <= 0:
        raise RuntimeError(f"{name} must be a positive integer")
    return value


_MAX_XLSX_ENTRIES = _positive_env_int("TLF_XLSX_MAX_ENTRIES", 1000)
_MAX_XLSX_ENTRY_BYTES = _positive_env_int("TLF_XLSX_MAX_ENTRY_BYTES", 16 * 1024 * 1024)
_MAX_XLSX_TOTAL_BYTES = _positive_env_int("TLF_XLSX_MAX_TOTAL_BYTES", 64 * 1024 * 1024)
_MAX_XLSX_COMPRESSION_RATIO = _positive_env_int("TLF_XLSX_MAX_COMPRESSION_RATIO", 100)
_XLSX_RATIO_CHECK_MIN_BYTES = _positive_env_int("TLF_XLSX_RATIO_CHECK_MIN_BYTES", 1024 * 1024)
_MAX_IMPORT_ROWS = _positive_env_int("TLF_XLSX_MAX_ROWS", 10_000)
_MAX_IMPORT_COLUMNS = _positive_env_int("TLF_XLSX_MAX_COLUMNS", 100)
_MAX_CELL_CHARS = _positive_env_int("TLF_XLSX_MAX_CELL_CHARS", 32_767)


def _preflight_xlsx(data: bytes) -> None:
    """Bound the OOXML container before openpyxl parses any workbook XML."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            names = [info.filename for info in infos]
            if len(infos) > _MAX_XLSX_ENTRIES:
                raise ValueError("Workbook contains too many archive entries.")
            if len(names) != len(set(names)):
                raise ValueError("Workbook contains duplicate archive entries.")
            if any(info.flag_bits & 0x1 for info in infos):
                raise ValueError("Encrypted workbooks are not supported.")
            if any(info.compress_type not in {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED} for info in infos):
                raise ValueError("Workbook uses an unsupported compression method.")
            if any(info.file_size > _MAX_XLSX_ENTRY_BYTES for info in infos):
                raise ValueError("Workbook archive entry is too large.")
            if sum(info.file_size for info in infos) > _MAX_XLSX_TOTAL_BYTES:
                raise ValueError("Workbook expands beyond the configured limit.")
            for info in infos:
                if info.file_size >= _XLSX_RATIO_CHECK_MIN_BYTES:
                    ratio = info.file_size / max(1, info.compress_size)
                    if ratio > _MAX_XLSX_COMPRESSION_RATIO:
                        raise ValueError("Workbook has an unsafe compression ratio.")
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required <= set(names):
                raise ValueError("File is not an Excel .xlsx workbook.")
            if archive.testzip() is not None:
                raise ValueError("Workbook archive failed its CRC check.")
    except zipfile.BadZipFile as exc:
        raise ValueError("File is not an Excel .xlsx workbook.") from exc


def _load_import_rows(data: bytes, preferred_sheet: str):
    _preflight_xlsx(data)
    try:
        wb = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False,
        )
    except Exception as exc:
        raise ValueError(f"Not a readable .xlsx file: {exc}") from exc
    try:
        if not wb.sheetnames:
            raise ValueError("The workbook has no sheets.")
        ws = wb[preferred_sheet] if preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        if ws.max_row > _MAX_IMPORT_ROWS or ws.max_column > _MAX_IMPORT_COLUMNS:
            raise ValueError(
                f"Workbook sheet exceeds the {_MAX_IMPORT_ROWS}-row / "
                f"{_MAX_IMPORT_COLUMNS}-column import limit."
            )
        rows = list(ws.iter_rows(values_only=True))
        for row_number, row in enumerate(rows, 1):
            for value in row:
                if isinstance(value, str) and len(value) > _MAX_CELL_CHARS:
                    raise ValueError(f"Workbook cell on row {row_number} is too long.")
        return rows, ws.title
    finally:
        wb.close()

# header alias  ->  finding field. Matched on a normalised header (lowercased, non-alphanumerics stripped), so "Output Label", "output_label", "OUTPUT" all hit "output".
_IMPORT_HEADERS = {
    "output": "output", "outputlabel": "output", "table": "output",
    "check": "check_id", "checkid": "check_id",
    "risk": "risk", "severity": "risk", "tier": "risk",
    "finding": "message", "message": "message", "issue": "message", "description": "message",
    "page": "page",
    "section": "section",
    "rowtype": "row_kind", "rowkind": "row_kind",
    "subjects": "subjects",
}


def _norm_hdr(s) -> str:
    return _re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _norm_lbl(s) -> str:
    return _re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _import_risk(v) -> str:
    """Any risk/severity spelling → the app's two tiers, High or Low."""
    return "High" if _TIER.get(str(v or "").strip().lower(), 1) == 0 else "Low"


def import_findings_xlsx(project_id: int, data: bytes, actor: str = "import") -> dict:
    """Insert findings from an .xlsx into `project_id`, ADDING to whatever is there (non-destructive). Returns a summary dict. Raises ValueError on an unreadable file or a sheet with no recognisable 'Finding' column."""
    # Prefer the export's own sheet; else the first sheet.
    rows, sheet_title = _load_import_rows(data, "AI Findings")
    if not rows:
        raise ValueError("The spreadsheet is empty.")

    # Find the header row: the first row (within the first few) that maps a "message" column — the export puts a title in row 1 and headers in row 2, but a hand-made sheet may start at row 1, so scan rather than assume.
    header_idx = col_map = None
    for i, row in enumerate(rows[:6]):
        m = {}
        for c, cell in enumerate(row):
            field = _IMPORT_HEADERS.get(_norm_hdr(cell))
            if field and field not in m:
                m[field] = c
        if "message" in m:
            header_idx, col_map = i, m
            break
    if col_map is None:
        raise ValueError("No 'Finding' (message) column found. Expected headers like "
                         "Output, Risk, Finding — matching the Excel export.")

    outputs = db.query("SELECT id, label FROM output WHERE project_id=?", (project_id,))
    by_label = {_norm_lbl(o["label"]): o for o in outputs}

    def cell(row, field):
        c = col_map.get(field)
        return row[c] if c is not None and c < len(row) else None

    run_id = db.insert("ai_run", project_id=project_id, kind="import",
                       started_at=db.now_iso(), finished_at=db.now_iso(), summary_json="{}")
    n = n_unmatched = n_deterministic = seq = 0
    for row in rows[header_idx + 1:]:
        if not row or not any(row):
            continue
        msg = cell(row, "message")
        if msg is None or str(msg).strip() == "":
            continue
        # Deterministic structural checks (FMT-*/XOUT-*) are generated automatically at project creation. A round-tripped export carries them too, so skip them here — the app's own copies are authoritative and re-importing must not duplicate them.
        if checks.is_deterministic_check(cell(row, "check_id")):
            n_deterministic += 1
            continue
        seq += 1
        lbl = cell(row, "output")
        out = by_label.get(_norm_lbl(lbl)) if lbl else None
        if lbl and not out:
            n_unmatched += 1
        risk = _import_risk(cell(row, "risk"))
        check_id = str(cell(row, "check_id") or f"IMP-{seq}").strip()
        page = cell(row, "page")
        try:
            page = int(page) if page is not None and str(page).strip() != "" else None
        except (TypeError, ValueError):
            page = None
        subj = cell(row, "subjects")
        subjects = [s.strip() for s in str(subj).split(",") if s.strip()] if subj else []
        db.insert("finding", project_id=project_id,
                  output_id=out["id"] if out else None, run_id=run_id,
                  check_id=check_id, severity=risk.lower(), risk=risk,
                  message=str(msg).strip(),
                  subjects=__import__("json").dumps(subjects), numbers="[]",
                  page=page, printed_page=None, pages_total=None,
                  section=(str(cell(row, "section")).strip() if cell(row, "section") else None),
                  row_kind=(str(cell(row, "row_kind")).strip() if cell(row, "row_kind") else None),
                  signature="", state="pending", badge="", phase="imported",
                  affected="[]" if out else __import__("json").dumps([str(lbl)] if lbl else []))
        n += 1

    # Importing an AI review counts as review completion: every output still 'Not Reviewed' that carries no finding of any kind (neither an imported AI finding nor a deterministic structural one) is a clean table and flips to Auto-approved — mirroring the auto-approve that follows a full in-app AI run. Only 'Not Reviewed' is overwritten, so a human-set status (or an already-approved table) is never lost.
    clean = db.query(
        "SELECT id FROM output WHERE project_id=? AND status='Not Reviewed' "
        "AND id NOT IN (SELECT output_id FROM finding "
        "WHERE project_id=? AND output_id IS NOT NULL)", (project_id, project_id))
    for row in clean:
        db.execute("UPDATE output SET status='Auto-approved' WHERE id=?", (row["id"],))
    n_auto = len(clean)
    if n_auto:
        db.audit(actor, "status.auto_approve", "project", project_id, project_id,
                 f"{n_auto} clean tables auto-approved on import")

    summary = {"imported": n, "unmatched_output": n_unmatched,
               "skipped_deterministic": n_deterministic, "auto_approved": n_auto,
               "sheet": sheet_title, "kind": "import"}
    db.execute("UPDATE ai_run SET summary_json=? WHERE id=?",
               (__import__("json").dumps(summary), run_id))
    return summary


# header alias -> comment field, matched on a normalised header (see _norm_hdr).
_COMMENT_IMPORT_HEADERS = {
    "id": "id", "commentid": "id",
    "table": "output", "output": "output", "outputlabel": "output",
    "comment": "body", "commenttext": "body",
    "replyto": "reply_to", "reply": "reply_to",
    "resolved": "resolved",
}


def _import_resolved(v) -> int:
    """A Resolved cell -> 0/1. Only '', 'yes', 'no' are valid (empty == no); the caller validates the spelling, so here 'yes' is the only truthy value."""
    return 1 if str(v or "").strip().lower() == "yes" else 0


def import_comments_xlsx(project_id: int, data: bytes, actor: str = "import") -> dict:
    """Re-import an edited comments sheet (the reverse of `comments_xlsx`).

Schema: ID | Table | Comment | Reply to | Resolved. (ID, Table) is the identity — a row whose (ID, Table) already exists REPLACES that comment (body / reply / resolved); a new (ID, Table) with an existing Table CREATES one. The whole sheet is validated up front and applied atomically: on ANY error nothing is written and a ValueError lists every problem.

Validation (see rules below): ID/Table/Comment non-empty; ID a positive integer; Table must exist; Resolved in {empty, yes, no} (case-insensitive); Reply to empty or a positive integer pointing to another comment in the SAME Table (existing or elsewhere in the sheet) and never itself; (ID, Table) unique within the sheet."""
    rows, sheet_title = _load_import_rows(data, "Review Comments")
    if not rows:
        raise ValueError("The spreadsheet is empty.")

    # Locate the header row: the first (within the first few) mapping id, output and body.
    header_idx = col_map = None
    for i, row in enumerate(rows[:6]):
        m = {}
        for c, cell_val in enumerate(row):
            field = _COMMENT_IMPORT_HEADERS.get(_norm_hdr(cell_val))
            if field and field not in m:
                m[field] = c
        if {"id", "output", "body"} <= set(m):
            header_idx, col_map = i, m
            break
    if col_map is None:
        raise ValueError("Missing the 'ID', 'Table' and 'Comment' columns. Expected the "
                         "headers from the comments export: ID, Table, Comment, Reply to, Resolved.")

    outputs = db.query("SELECT id, label FROM output WHERE project_id=?", (project_id,))
    by_label = {_norm_lbl(o["label"]): o for o in outputs}
    has_resolved = "resolved" in col_map

    def cell(row, field):
        c = col_map.get(field)
        return row[c] if c is not None and c < len(row) else None

    def _pos_int(v):
        """Return a positive int, or None if the cell isn't one (empty, 0, negative, float)."""
        s = str(v).strip() if v is not None else ""
        if s == "" or not _re.fullmatch(r"\d+", s):
            return None
        n = int(s)
        return n if n > 0 else None

    # ---- Pass 1: parse + validate every data row, collecting errors -------------------- #
    errors: list[str] = []
    parsed = []                 # (rownum, oid, num, body, reply_to, resolved)
    seen_keys: dict = {}        # (oid, num) -> first row number (dup detection)
    sheet_ids: dict = {}        # oid -> set(num) defined in this sheet
    for offset, row in enumerate(rows[header_idx + 1:]):
        rownum = header_idx + 2 + offset            # 1-based row in the sheet
        if not row or not any(v not in (None, "") for v in row):
            continue

        raw_id, raw_tbl = cell(row, "id"), cell(row, "output")
        body = cell(row, "body")
        body = str(body).strip() if body is not None else ""
        raw_reply, raw_res = cell(row, "reply_to"), (cell(row, "resolved") if has_resolved else None)

        lbl = str(raw_tbl).strip() if raw_tbl is not None else ""
        # Guard against `raw_id or ""` — a literal 0 is falsy but NOT empty; it must fall through to the positive-integer check below, not be reported as a blank cell.
        id_str = str(raw_id).strip() if raw_id is not None else ""
        if not id_str or not lbl or not body:
            errors.append(f"Row {rownum}: ID, Table and Comment must all be filled in.")
            continue

        num = _pos_int(raw_id)
        if num is None:
            errors.append(f"Row {rownum}: ID must be a positive whole number (got {raw_id!r}).")
            continue

        out = by_label.get(_norm_lbl(lbl))
        if out is None:
            errors.append(f"Row {rownum}: Table {lbl!r} does not exist in this project.")
            continue
        oid = out["id"]

        res_str = str(raw_res or "").strip().lower()
        if res_str not in ("", "yes", "no"):
            errors.append(f"Row {rownum}: Resolved must be empty, 'Yes' or 'No' (got {raw_res!r}).")
            continue

        reply_to = None
        if raw_reply is not None and str(raw_reply).strip() != "":
            reply_to = _pos_int(raw_reply)
            if reply_to is None:
                errors.append(f"Row {rownum}: 'Reply to' must be empty or a positive whole "
                              f"number (got {raw_reply!r}).")
                continue
            if reply_to == num:
                errors.append(f"Row {rownum}: comment {num} in {lbl!r} cannot reply to itself.")
                continue

        key = (oid, num)
        if key in seen_keys:
            errors.append(f"Row {rownum}: duplicate (ID {num}, Table {lbl!r}) — also on "
                          f"row {seen_keys[key]}.")
            continue
        seen_keys[key] = rownum
        sheet_ids.setdefault(oid, set()).add(num)
        parsed.append((rownum, oid, num, body, reply_to,
                       _import_resolved(res_str) if has_resolved else 0))

    # Reply targets must resolve within the same Table (existing comments ∪ this sheet).
    existing_nums: dict = {}
    for oid in {p[1] for p in parsed}:
        existing_nums[oid] = {r["num"] for r in db.query(
            "SELECT num FROM comment WHERE output_id=? AND num IS NOT NULL", (oid,))}
    for rownum, oid, num, body, reply_to, _res in parsed:
        if reply_to is not None and reply_to not in (existing_nums.get(oid, set()) | sheet_ids.get(oid, set())):
            errors.append(f"Row {rownum}: 'Reply to' {reply_to} does not match any comment "
                          f"in this Table.")

    if errors:
        raise ValueError("Import failed — fix these and re-import:\n- " + "\n- ".join(errors))

    # ---- Pass 2 (atomic): upsert bodies, then wire up replies -------------------------- #
    gid: dict = {}              # (oid, num) -> global comment id
    updated = created = 0
    for _rownum, oid, num, body, _reply_to, resolved in parsed:
        existing = db.one("SELECT id FROM comment WHERE output_id=? AND num=?", (oid, num))
        if existing:
            db.execute("UPDATE comment SET body=?, resolved=? WHERE id=?",
                       (body, resolved, existing["id"]))
            gid[(oid, num)] = existing["id"]
            updated += 1
        else:
            cid = db.insert("comment", project_id=project_id, output_id=oid, num=num,
                            title="", body=body, source="manual", author=actor,
                            parent_id=None, resolved=resolved, created_at=db.now_iso())
            gid[(oid, num)] = cid
            # A new comment nudges the output into review, matching add_comment.
            db.execute("UPDATE output SET status='In Progress' "
                       "WHERE id=? AND status='Not Reviewed'", (oid,))
            created += 1

    for _rownum, oid, num, _body, reply_to, _res in parsed:
        parent = gid.get((oid, reply_to)) if reply_to is not None else None
        if parent is None and reply_to is not None:
            # Target is an existing comment not itself in the sheet — look it up.
            row = db.one("SELECT id FROM comment WHERE output_id=? AND num=?", (oid, reply_to))
            parent = row["id"] if row else None
        db.execute("UPDATE comment SET parent_id=? WHERE id=?", (parent, gid[(oid, num)]))

    return {"updated": updated, "created": created, "sheet": sheet_title}
